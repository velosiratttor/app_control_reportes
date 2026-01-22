import streamlit as st
import pandas as pd
from openpyxl.styles import Font,Alignment,PatternFill,Border,Side

ruta='reporte.xlsx'

st.title("registro de actividades diarias")
st.write("app de registro de actividades")

if "datos" not in st.session_state:
    st.session_state.datos=[]


nombre= st.text_input("Nombre del analista")
fecha=st.date_input("fecha de la actividad")
actividad=st.selectbox(
    "Seleccione su actividad",
    ["AUTORIZACIÓN PARA EL INGRESO DE CALIFICACIONES (AÑOS ANTERIORES)",
     "REENVIO DE CERTIFICADOS DE CURSOS VIRTUALES",
     "CORRECIÓN DE DATOS DE ESTUDIANTE/PROTAGONISTA",
     "ASISTENCIA VIA TELEFONICA A C.T.",
     "SOLICITUD PARA CAMBIO DE CORREO"],
     index=None,
     placeholder="Debes seleccionar tus acticvidades realizadas",
     accept_new_options=True,
)
estado=st.selectbox(
    "Estado de la actividad",
    ["ATENDIDO",
     "EN ATENCION",
     "ESCALADO"],
     index=None,
     placeholder="seleccione el estado de tu actividad",
     accept_new_options=True,
)


if st.button("agregar"):
    st.session_state.datos.append({
        "Nombre":nombre,
        "Fecha":fecha,
        "Actividad":actividad,
        "Estado":estado
    })

df= pd.DataFrame(st.session_state.datos)
st.dataframe(df)

if not df.empty:
    if st.button("Descargar excel con encabezados"):
        ruta="reporte.xlsx"
        with pd.ExcelWriter(ruta,engine="openpyxl") as writer:
            df.to_excel(
                writer,
                sheet_name="Reporte",
                index=False,
                startrow=5
            )
            hoja=writer.sheets["Reporte"]

            hoja.merge_cells("A1:B1")
            hoja.merge_cells("A2:B2")
            hoja.merge_cells("A3:B3")

            hoja["A1"]="INSTITUTO NACIONAL TÉCNICO Y TECNOLÓGICO"
            hoja["A2"] = "ATENCIÓN DE SOLICITUDES"
            hoja["A3"] = "DIRECCIÓN DE FORMACIÓN PROFESIONAL"

            for fila in range(1, 4):
                celda = hoja[f"A{fila}"]
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center")
#st.write("Nombre",nombre)
#st.write("Fecha",fecha)
#st.write("Actividad",actividad)
#st.write("Estado",estado)